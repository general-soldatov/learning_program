import openpyxl
import os
import requests
import yaml
import jinja2
from docxtpl import DocxTemplate
from progress.bar import IncrementalBar
from abc import ABC, abstractmethod
from bs4 import BeautifulSoup
from .config import Program

def create_division(message: str) -> str:
    columns = os.get_terminal_size().columns
    padding_length = columns - len(message) - 2
    left_padding = '#' * (padding_length // 2)
    right_padding = '#' * ((padding_length + 1) // 2)
    return f"{left_padding} {message} {right_padding}"

class Parser(ABC):
    def __init__(self, url=None):
        self.__get_url(url)
        self.parser = "html.parser"
        self.attr = {}

    def __get_url(self, url):
        if url:
            self.url = url
            self.page = requests.get(self.url).text

    def _find_data(self, all_=False, **kwargs) -> list | str:
        self._soup = BeautifulSoup(self.page, self.parser)
        if not all_:
            return self._soup.find(**self.attr, **kwargs)
        return self._soup.find_all(**self.attr, **kwargs)

    @abstractmethod
    def find_data(self, *args, **kwargs):
        pass

class ParseXls:
    def __init__(self, path, name=None, max_RC: bool|tuple = False, _bar_max=5):
        self.bar = IncrementalBar("Load Data with Exel", max = _bar_max, suffix='%(percent)d%%')
        self.workbook = openpyxl.load_workbook(path, read_only=True)
        self.bar.next()
        self._worksheet(name)
        self.row_index = None
        if self.sheet.max_row and not max_RC:
            self.max_row = self.sheet.max_row
            self.max_col = self.sheet.max_column
        self.max_row, self.max_col = max_RC
        self.bar.next()

    def _worksheet(self, name):
        self.sheet = self.workbook.active
        if name:
            self.sheet = self.workbook[name]

    def _search_to_column(self, to_scan: str, col: int, start: int = 1) -> int:
        for row in range(start, self.max_row + 1):
            if self.sheet.cell(row, col).value == to_scan:
                return row

    @staticmethod
    def _search_with_column(sheet, to_scan: str, col: int, start: int = 1, stop: int = 1) -> int:
        for row in range(start, stop + 1):
            if sheet.cell(row, col).value == to_scan:
                return row


class YamlCreator:
    def __init__(self, path: str, encoding: str = 'utf-8', _bar_max=5):
        self.bar = IncrementalBar("Loading Project", max = _bar_max, suffix='%(percent)d%%')
        self.bar.next()
        self._path = path
        with open(path, 'r', encoding=encoding) as project:
            self.bar.next()
            self._data = yaml.safe_load(project)
            self.bar.next()

    def recording(self, path: str = None, encoding: str = 'utf-8'):
        if not path:
            path = self._path
        self.bar.next()
        with open(path, 'w', encoding=encoding) as project:
            yaml.dump(self._data, project)
            self.bar.next()
            self.bar.finish()


class WordDocument:
    def __init__(self, data: Program, path_doc: str = None, path: str = "", _bar_max=5):
        self.bar = IncrementalBar("Create Word document", max = _bar_max, suffix='%(percent)d%%')
        self.bar.next()
        self.doc: DocxTemplate = DocxTemplate(path_doc)
        self.bar.next()
        self.data: Program = data
        self.type_doc: str = "NoneType"
        self.name_file = ""
        self.path_doc = path_doc
        self.jinja_env = jinja2.Environment()
        self.context = {**data.__dict__, 'name': 'NoName'}
        self.path = path

    def create_dir(self):
        self.path_dir = self.path + f"/{self.data.code} {self.context['name']}"
        os.makedirs(self.path_dir, exist_ok=True)
        self.bar.next()

    def create_name(self):
        self.create_dir()
        self.name_file = f"{self.path_dir}/{self.type_doc} {self.data.code} {self.context['name']}.docx"
        self.bar.next()

    def create_document(self) -> str | Exception:
        self.create_name()
        info = self.__document_create(self.doc, self.name_file, self.context, self.jinja_env)
        self.bar.finish()
        return f"Word document has been created and placed on the path\n{info}"


    @staticmethod
    def __document_create(docs: DocxTemplate, name_file: str, context: dict, jinja_env: jinja2.Environment) -> str | Exception:
        try:
            docs.render(context, jinja_env)
            docs.save(name_file)
            return name_file
        except Exception as e:
            return e